"""Prova, ponta a ponta, a garantia de compliance MAIS crítica do lado francês do
projeto (ver CLAUDE.md — "registros 'diffusion partielle' NUNCA entram em lista de
prospecção", filtro hard, sem exceção):

    stock SIRENE com statut_diffusion "P" (diffusion partielle)
        -> map_unite_legale_etablissement_to_canonical: flag_difusao_restrita=True
        -> segmentation.suppression: is_hard_excluded=True, apply_suppression_gate exclui
        -> export.exporters.export_csv/export_xlsx: NUNCA aparece no arquivo gerado

Cada etapa é verificada isoladamente e depois em conjunto, com um lead "O" (aberto)
irmão sempre presente ao lado para provar que a exclusão é seletiva (só o restrito
some), não um efeito colateral de todos os leads sumirem.
"""

from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import load_workbook

from src.etl.canonical import map_unite_legale_etablissement_to_canonical
from src.export.exporters import export_csv, export_xlsx
from src.segmentation.suppression import (
    SuppressionList,
    apply_suppression_gate,
    is_hard_excluded,
)

NO_SUPPRESSION = SuppressionList()


def _unite_legale(*, siren: str, statut_diffusion: str | None) -> dict[str, object]:
    return {
        "entidade": "UNITE_LEGALE",
        "siren": siren,
        "statut_diffusion": statut_diffusion,
        "flag_difusao_restrita": (statut_diffusion or "").strip().upper() != "O",
        "situacao": "A",
        "razao_social": f"EMPRESA {siren}",
        "nome_fantasia": None,
        "sigla": None,
        "natureza_juridica": "5710",
        "cod_atividade": "62.01Z",
        "categoria_empresa": "PME",
        "data_criacao": None,
    }


def _etablissement(
    *, siren: str, siret: str, statut_diffusion: str | None = "O"
) -> dict[str, object]:
    return {
        "entidade": "ETABLISSEMENT",
        "siren": siren,
        "siret": siret,
        "statut_diffusion": statut_diffusion,
        "flag_difusao_restrita": (statut_diffusion or "").strip().upper() != "O",
        "situacao": "A",
        "nome_fantasia": None,
        "cod_atividade": "62.01Z",
        "municipio": "PARIS",
        "cep": "75001",
        "data_criacao": None,
    }


# Par restrito: unidade legal em "diffusion partielle" ("P"), estabelecimento aberto
# ("O") -- prova que basta UM dos dois lados estar restrito (ver docstring de
# map_unite_legale_etablissement_to_canonical: é sempre o OR dos dois).
RESTRICTED_SIREN = "999888777"
RESTRICTED_SIRET = "99988877700010"
RESTRICTED_UNITE_LEGALE = _unite_legale(siren=RESTRICTED_SIREN, statut_diffusion="P")
RESTRICTED_ETABLISSEMENT = _etablissement(
    siren=RESTRICTED_SIREN, siret=RESTRICTED_SIRET, statut_diffusion="O"
)

# Par irmão, totalmente aberto -- prova que a exclusão é seletiva.
OPEN_SIREN = "111222333"
OPEN_SIRET = "11122233300010"
OPEN_UNITE_LEGALE = _unite_legale(siren=OPEN_SIREN, statut_diffusion="O")
OPEN_ETABLISSEMENT = _etablissement(siren=OPEN_SIREN, siret=OPEN_SIRET, statut_diffusion="O")


class TestMappingSetsFlag:
    """Etapa 1: statut_diffusion "P" (em qualquer um dos dois lados) -> flag=True."""

    def test_diffusion_partielle_sets_flag_true(self) -> None:
        lead = map_unite_legale_etablissement_to_canonical(
            RESTRICTED_UNITE_LEGALE, RESTRICTED_ETABLISSEMENT
        )
        assert lead["flag_difusao_restrita"] is True
        assert lead["id_estab"] == RESTRICTED_SIRET

    def test_open_pair_sets_flag_false(self) -> None:
        lead = map_unite_legale_etablissement_to_canonical(OPEN_UNITE_LEGALE, OPEN_ETABLISSEMENT)
        assert lead["flag_difusao_restrita"] is False

    def test_flag_true_even_when_only_etablissement_is_restricted(self) -> None:
        """O oposto do caso principal: unidade legal aberta, só o estabelecimento em
        diffusion partielle -- também deve virar True (OR, não "só a unidade legal
        conta")."""
        unite_legale = _unite_legale(siren=OPEN_SIREN, statut_diffusion="O")
        etablissement = _etablissement(siren=OPEN_SIREN, siret=OPEN_SIRET, statut_diffusion="P")
        lead = map_unite_legale_etablissement_to_canonical(unite_legale, etablissement)
        assert lead["flag_difusao_restrita"] is True


class TestSuppressionExcludesFlaggedLead:
    """Etapa 2: flag=True -> is_hard_excluded=True -> apply_suppression_gate remove."""

    def test_is_hard_excluded(self) -> None:
        lead = map_unite_legale_etablissement_to_canonical(
            RESTRICTED_UNITE_LEGALE, RESTRICTED_ETABLISSEMENT
        )
        assert is_hard_excluded(lead) is True

    def test_open_lead_is_not_hard_excluded(self) -> None:
        lead = map_unite_legale_etablissement_to_canonical(OPEN_UNITE_LEGALE, OPEN_ETABLISSEMENT)
        assert is_hard_excluded(lead) is False

    def test_apply_suppression_gate_removes_only_the_restricted_lead(self) -> None:
        restricted_lead = map_unite_legale_etablissement_to_canonical(
            RESTRICTED_UNITE_LEGALE, RESTRICTED_ETABLISSEMENT
        )
        open_lead = map_unite_legale_etablissement_to_canonical(
            OPEN_UNITE_LEGALE, OPEN_ETABLISSEMENT
        )

        final, report = apply_suppression_gate([restricted_lead, open_lead], NO_SUPPRESSION)

        assert [lead["id_estab"] for lead in final] == [OPEN_SIRET]
        assert report.n_in == 2
        assert report.n_hard_excluded == 1
        assert report.n_out == 1


class TestExportNeverContainsFlaggedLead:
    """Etapa 3: nem export_csv nem export_xlsx deixam o registro restrito chegar ao
    arquivo final — "excluído pela supressão em QUALQUER exportação"."""

    def test_export_csv_excludes_diffusion_partielle_record(self, tmp_path: Path) -> None:
        restricted_lead = map_unite_legale_etablissement_to_canonical(
            RESTRICTED_UNITE_LEGALE, RESTRICTED_ETABLISSEMENT
        )
        open_lead = map_unite_legale_etablissement_to_canonical(
            OPEN_UNITE_LEGALE, OPEN_ETABLISSEMENT
        )
        dest = tmp_path / "leads.csv"

        result = export_csv(
            [restricted_lead, open_lead],
            dest,
            suppression=NO_SUPPRESSION,
            audit_log_path=tmp_path / "audit.parquet",
        )

        assert result.n_exported == 1
        assert result.suppression_report.n_hard_excluded == 1

        with dest.open(encoding="utf-8", newline="") as f:
            ids_no_arquivo = [row["id_estab"] for row in csv.DictReader(f)]
        assert RESTRICTED_SIRET not in ids_no_arquivo
        assert ids_no_arquivo == [OPEN_SIRET]

    def test_export_xlsx_excludes_diffusion_partielle_record(self, tmp_path: Path) -> None:
        restricted_lead = map_unite_legale_etablissement_to_canonical(
            RESTRICTED_UNITE_LEGALE, RESTRICTED_ETABLISSEMENT
        )
        open_lead = map_unite_legale_etablissement_to_canonical(
            OPEN_UNITE_LEGALE, OPEN_ETABLISSEMENT
        )
        dest = tmp_path / "leads.xlsx"

        result = export_xlsx(
            [restricted_lead, open_lead],
            dest,
            suppression=NO_SUPPRESSION,
            audit_log_path=tmp_path / "audit.parquet",
        )

        assert result.n_exported == 1

        workbook = load_workbook(dest)
        sheet = workbook.active
        assert sheet is not None
        id_estab_col = next(
            i for i, cell in enumerate(next(sheet.iter_rows(max_row=1))) if cell.value == "id_estab"
        )
        ids_na_planilha = [row[id_estab_col].value for row in sheet.iter_rows(min_row=2)]
        assert RESTRICTED_SIRET not in ids_na_planilha
        assert ids_na_planilha == [OPEN_SIRET]

    def test_export_with_only_the_restricted_lead_produces_empty_export(
        self, tmp_path: Path
    ) -> None:
        """Caso limite: se o restrito é o ÚNICO lead candidato, a exportação deve
        sair vazia -- nunca "melhor exportar algo do que nada"."""
        restricted_lead = map_unite_legale_etablissement_to_canonical(
            RESTRICTED_UNITE_LEGALE, RESTRICTED_ETABLISSEMENT
        )
        dest = tmp_path / "leads.csv"

        result = export_csv(
            [restricted_lead],
            dest,
            suppression=NO_SUPPRESSION,
            audit_log_path=tmp_path / "audit.parquet",
        )

        assert result.n_exported == 0
        with dest.open(encoding="utf-8", newline="") as f:
            rows = list(csv.DictReader(f))
        assert rows == []
