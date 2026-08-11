"""Testes de `export/exporters.py`: o gate de supressão (síntese + difusão restrita
nunca aparecem no arquivo exportado), o registro em audit_log, e o conteúdo dos
arquivos CSV/Excel gerados.
"""

from __future__ import annotations

import csv
import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from src.compliance.audit_log import read_audit_log
from src.export.exporters import EXPORT_COLUMNS, ExportError, export_csv, export_txt, export_xlsx
from src.segmentation.suppression import SuppressionList


def _lead(**overrides: object) -> dict[str, object]:
    base: dict[str, object] = {
        "pais": "BR",
        "id_legal": "11111111",
        "id_estab": "11111111000191",
        "razao_social": "EMPRESA TESTE LTDA",
        "nome_fantasia": "EMPRESA TESTE",
        "cod_atividade": "8630501",
        "situacao": "ATIVA",
        "regiao": "SP",
        "municipio": "SAO PAULO",
        "cep": "01310100",
        "telefone": "01122334455",
        "email": "contato@empresateste.com.br",
        "data_inicio_atividade": None,
        "porte": "MICRO EMPRESA",
        "capital_social": None,
        "natureza_juridica": "SOCIEDADE EMPRESARIA LIMITADA",
        "score_icp": 80.0,
        "fonte": "BR_RECEITA",
        "enriquecido_em": None,
        "is_synthetic": False,
        "flag_difusao_restrita": False,
    }
    return {**base, **overrides}


NO_SUPPRESSION = SuppressionList()


def _txt_field(text: str, label: str) -> str:
    """Valor de um campo `Rótulo: valor` no texto exportado -- não usa a string
    inteira porque `_write_txt` alinha os rótulos (padding variável entre `:` e o
    valor, ver export/exporters._write_txt)."""
    for line in text.splitlines():
        if line.startswith(f"{label}:"):
            return line.split(":", 1)[1].strip()
    raise AssertionError(f"campo {label!r} não encontrado no texto exportado")


class TestSuppressionGateOnExportCsv:
    def test_synthetic_lead_never_reaches_the_file(self, tmp_path: Path) -> None:
        leads = [_lead(id_estab="ok"), _lead(id_estab="synth", is_synthetic=True)]
        dest = tmp_path / "out.csv"

        result = export_csv(
            leads, dest, suppression=NO_SUPPRESSION, audit_log_path=tmp_path / "audit.parquet"
        )

        assert result.n_exported == 1
        with dest.open(encoding="utf-8", newline="") as f:
            ids = [row["id_estab"] for row in csv.DictReader(f)]
        assert ids == ["ok"]

    def test_diffusion_partielle_lead_never_reaches_the_file(self, tmp_path: Path) -> None:
        leads = [_lead(id_estab="ok"), _lead(id_estab="fr-restrito", flag_difusao_restrita=True)]
        dest = tmp_path / "out.csv"

        result = export_csv(
            leads, dest, suppression=NO_SUPPRESSION, audit_log_path=tmp_path / "audit.parquet"
        )

        assert result.n_exported == 1
        with dest.open(encoding="utf-8", newline="") as f:
            ids = [row["id_estab"] for row in csv.DictReader(f)]
        assert ids == ["ok"]

    def test_duplicate_id_estab_deduplicated(self, tmp_path: Path) -> None:
        leads = [_lead(id_estab="A"), _lead(id_estab="A")]
        dest = tmp_path / "out.csv"

        result = export_csv(
            leads, dest, suppression=NO_SUPPRESSION, audit_log_path=tmp_path / "audit.parquet"
        )

        assert result.n_exported == 1

    def test_suppressed_lead_removed(self, tmp_path: Path) -> None:
        leads = [
            _lead(id_estab="A", email="a@empresaa.com"),
            _lead(id_estab="B", email="b@empresab.com"),
        ]
        suppression = SuppressionList(ids_estab=frozenset({"A"}))
        dest = tmp_path / "out.csv"

        result = export_csv(
            leads, dest, suppression=suppression, audit_log_path=tmp_path / "audit.parquet"
        )

        assert result.n_exported == 1
        with dest.open(encoding="utf-8", newline="") as f:
            ids = [row["id_estab"] for row in csv.DictReader(f)]
        assert ids == ["B"]

    def test_suppression_report_reflects_every_rule(self, tmp_path: Path) -> None:
        leads = [
            _lead(id_estab="ok", email="ok@empresa1.com"),
            _lead(id_estab="synth", is_synthetic=True, email="synth@empresa2.com"),
            _lead(id_estab="fr", flag_difusao_restrita=True, email="fr@empresa3.com"),
            _lead(id_estab="dup", email="dup@empresa4.com"),
            _lead(id_estab="dup", email="dup@empresa4.com"),
            _lead(id_estab="suprimido", email="suprimido@empresa5.com"),
        ]
        suppression = SuppressionList(ids_estab=frozenset({"suprimido"}))
        dest = tmp_path / "out.csv"

        result = export_csv(
            leads, dest, suppression=suppression, audit_log_path=tmp_path / "audit.parquet"
        )

        report = result.suppression_report
        assert report.n_in == 6
        assert report.n_hard_excluded == 2
        assert report.n_deduped_id_estab == 1
        assert report.n_suppressed == 1
        assert report.n_out == result.n_exported == 2


class TestExportCsvContent:
    def test_header_and_column_order(self, tmp_path: Path) -> None:
        dest = tmp_path / "out.csv"
        export_csv(
            [_lead()], dest, suppression=NO_SUPPRESSION, audit_log_path=tmp_path / "audit.parquet"
        )

        with dest.open(encoding="utf-8", newline="") as f:
            header = next(csv.reader(f))
        assert tuple(header) == EXPORT_COLUMNS

    def test_excludes_compliance_flags_from_columns(self) -> None:
        assert "is_synthetic" not in EXPORT_COLUMNS
        assert "flag_difusao_restrita" not in EXPORT_COLUMNS

    def test_row_values(self, tmp_path: Path) -> None:
        dest = tmp_path / "out.csv"
        export_csv(
            [_lead(razao_social="ACME LTDA")],
            dest,
            suppression=NO_SUPPRESSION,
            audit_log_path=tmp_path / "audit.parquet",
        )

        with dest.open(encoding="utf-8", newline="") as f:
            row = next(csv.DictReader(f))
        assert row["razao_social"] == "ACME LTDA"
        assert row["fonte"] == "BR_RECEITA"

    def test_empty_after_suppression_still_writes_header_only(self, tmp_path: Path) -> None:
        dest = tmp_path / "out.csv"
        result = export_csv(
            [_lead(is_synthetic=True)],
            dest,
            suppression=NO_SUPPRESSION,
            audit_log_path=tmp_path / "audit.parquet",
        )

        assert result.n_exported == 0
        with dest.open(encoding="utf-8", newline="") as f:
            rows = list(csv.reader(f))
        assert len(rows) == 1  # só o cabeçalho

    def test_no_columns_raises_export_error(self, tmp_path: Path) -> None:
        with pytest.raises(ExportError):
            export_csv(
                [_lead()],
                tmp_path / "out.csv",
                suppression=NO_SUPPRESSION,
                columns=[],
                audit_log_path=tmp_path / "audit.parquet",
            )


class TestExportXlsxContent:
    def test_header_and_rows(self, tmp_path: Path) -> None:
        dest = tmp_path / "out.xlsx"
        export_xlsx(
            [_lead(razao_social="ACME LTDA")],
            dest,
            suppression=NO_SUPPRESSION,
            audit_log_path=tmp_path / "audit.parquet",
        )

        wb = load_workbook(dest)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        assert rows[0] == EXPORT_COLUMNS
        assert rows[1][EXPORT_COLUMNS.index("razao_social")] == "ACME LTDA"

    def test_gate_applies_to_xlsx_too(self, tmp_path: Path) -> None:
        leads = [_lead(id_estab="ok"), _lead(id_estab="synth", is_synthetic=True)]
        dest = tmp_path / "out.xlsx"

        result = export_xlsx(
            leads, dest, suppression=NO_SUPPRESSION, audit_log_path=tmp_path / "audit.parquet"
        )

        assert result.n_exported == 1
        wb = load_workbook(dest)
        rows = list(wb.active.iter_rows(values_only=True))
        assert len(rows) == 2  # cabeçalho + 1 lead


class TestExportTxtContent:
    def test_gate_applies_to_txt_too(self, tmp_path: Path) -> None:
        leads = [_lead(id_estab="ok"), _lead(id_estab="synth", is_synthetic=True)]
        dest = tmp_path / "out.txt"

        result = export_txt(
            leads, dest, suppression=NO_SUPPRESSION, audit_log_path=tmp_path / "audit.parquet"
        )

        assert result.n_exported == 1
        text = dest.read_text(encoding="utf-8")
        assert text.count("=" * 80) == 2  # uma divisória de abertura + uma de fechamento, 1 lead
        assert "synth" not in text

    def test_block_contains_every_column_as_labeled_line(self, tmp_path: Path) -> None:
        dest = tmp_path / "out.txt"
        export_txt(
            [_lead(razao_social="ACME LTDA")],
            dest,
            suppression=NO_SUPPRESSION,
            audit_log_path=tmp_path / "audit.parquet",
        )

        text = dest.read_text(encoding="utf-8")
        assert "ACME LTDA" in text
        assert _txt_field(text, "Razão social") == "ACME LTDA"
        assert _txt_field(text, "CNPJ / SIRET") == "11111111000191"
        assert _txt_field(text, "E-mail") == "contato@empresateste.com.br"
        assert _txt_field(text, "Fonte") == "BR_RECEITA"

    def test_missing_value_rendered_as_dash(self, tmp_path: Path) -> None:
        dest = tmp_path / "out.txt"
        export_txt(
            [_lead(nome_fantasia=None)],
            dest,
            suppression=NO_SUPPRESSION,
            audit_log_path=tmp_path / "audit.parquet",
        )

        text = dest.read_text(encoding="utf-8")
        assert _txt_field(text, "Nome fantasia") == "—"

    def test_multiple_leads_get_separate_blocks(self, tmp_path: Path) -> None:
        # E-mails de domínios diferentes -- domínio igual dispara
        # dedupe_by_email_domain (ver segmentation/suppression.py) e colapsaria os
        # dois leads em um só, o que não é o que este teste quer exercitar.
        dest = tmp_path / "out.txt"
        export_txt(
            [
                _lead(id_estab="A", razao_social="EMPRESA A", email="a@empresaa.com"),
                _lead(id_estab="B", razao_social="EMPRESA B", email="b@empresab.com"),
            ],
            dest,
            suppression=NO_SUPPRESSION,
            audit_log_path=tmp_path / "audit.parquet",
        )

        text = dest.read_text(encoding="utf-8")
        assert text.count("=" * 80) == 4  # 2 divisórias por lead, 2 leads
        assert "EMPRESA A" in text
        assert "EMPRESA B" in text

    def test_empty_after_suppression_still_writes_a_file(self, tmp_path: Path) -> None:
        dest = tmp_path / "out.txt"
        result = export_txt(
            [_lead(is_synthetic=True)],
            dest,
            suppression=NO_SUPPRESSION,
            audit_log_path=tmp_path / "audit.parquet",
        )

        assert result.n_exported == 0
        assert dest.exists()
        assert "=" not in dest.read_text(encoding="utf-8")

    def test_no_columns_raises_export_error(self, tmp_path: Path) -> None:
        with pytest.raises(ExportError):
            export_txt(
                [_lead()],
                tmp_path / "out.txt",
                suppression=NO_SUPPRESSION,
                columns=[],
                audit_log_path=tmp_path / "audit.parquet",
            )


class TestAuditLogIntegration:
    def test_export_records_audit_event_before_returning(self, tmp_path: Path) -> None:
        dest = tmp_path / "out.csv"
        audit_path = tmp_path / "audit.parquet"

        export_csv(
            [_lead(id_estab="A")],
            dest,
            suppression=NO_SUPPRESSION,
            filtros={"regiao": "SP"},
            usuario="italo",
            audit_log_path=audit_path,
        )

        log = read_audit_log(audit_path)
        assert log.height == 1
        row = log.to_dicts()[0]
        assert row["operacao"] == "export_csv"
        assert row["usuario"] == "italo"
        assert row["n_registros"] == 1
        assert json.loads(row["filtros"]) == {"regiao": "SP"}
        assert row["destino"] == str(dest)

    def test_audit_n_registros_reflects_post_gate_count(self, tmp_path: Path) -> None:
        leads = [_lead(id_estab="ok"), _lead(id_estab="synth", is_synthetic=True)]
        audit_path = tmp_path / "audit.parquet"

        export_csv(
            leads, tmp_path / "out.csv", suppression=NO_SUPPRESSION, audit_log_path=audit_path
        )

        row = read_audit_log(audit_path).to_dicts()[0]
        assert row["n_registros"] == 1  # não 2 -- o sintético já foi barrado

    def test_xlsx_export_uses_distinct_operacao(self, tmp_path: Path) -> None:
        audit_path = tmp_path / "audit.parquet"
        export_xlsx(
            [_lead()], tmp_path / "out.xlsx", suppression=NO_SUPPRESSION, audit_log_path=audit_path
        )

        row = read_audit_log(audit_path).to_dicts()[0]
        assert row["operacao"] == "export_xlsx"

    def test_txt_export_uses_distinct_operacao(self, tmp_path: Path) -> None:
        audit_path = tmp_path / "audit.parquet"
        export_txt(
            [_lead()], tmp_path / "out.txt", suppression=NO_SUPPRESSION, audit_log_path=audit_path
        )

        row = read_audit_log(audit_path).to_dicts()[0]
        assert row["operacao"] == "export_txt"

    def test_multiple_exports_accumulate_in_the_log(self, tmp_path: Path) -> None:
        audit_path = tmp_path / "audit.parquet"
        kwargs = {"suppression": NO_SUPPRESSION, "audit_log_path": audit_path}
        export_csv([_lead()], tmp_path / "a.csv", **kwargs)
        export_csv([_lead()], tmp_path / "b.csv", **kwargs)

        assert read_audit_log(audit_path).height == 2

    def test_audit_event_recorded_even_when_result_is_empty(self, tmp_path: Path) -> None:
        audit_path = tmp_path / "audit.parquet"
        export_csv(
            [_lead(is_synthetic=True)],
            tmp_path / "out.csv",
            suppression=NO_SUPPRESSION,
            audit_log_path=audit_path,
        )

        row = read_audit_log(audit_path).to_dicts()[0]
        assert row["n_registros"] == 0
